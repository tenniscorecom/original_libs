"""
excel/sheet.py — ワークシートの高レベルラッパー

ExcelFile.sheet() から取得し、セル書き込み・行書き込み・列幅調整・
ヘッダー固定などをシート単位で行う（sheet_name を毎回渡さなくてよい）。

使い方:
    from comken.excel import ExcelFile

    # 既存ファイルを開いて書き込む
    with ExcelFile("report.xlsx") as f:
        s = f.sheet("Sheet1")
        s["A1"] = "売上レポート"          # セル参照で書き込み
        title = s["A1"]                   # セル参照で読み取り
        s.write_row(3, ["日付", "金額"])  # 3行目に横並びで書く
        s.append_row(["2026-07-12", 1000])  # 最終行の下に追記
        s.auto_width()                    # 列幅を内容に合わせる（日本語対応）
        s.freeze_header()                 # 1行目を固定
        f.save()

    # 新規ブックを作ってレポートを出力する
    rows = CsvReader("data.csv").rows()
    with ExcelFile.create(r"C:\\作業\\report.xlsx") as f:
        s = f.sheet("Sheet1")
        s.write_table(rows)               # ヘッダー行 + データ行をまとめて書く
        s.auto_width()
        s.freeze_header()
        f.save()
"""

from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


def _display_width(value) -> int:
    """列幅計算用の表示幅を返す（全角文字は2文字ぶんとして数える）。"""
    return sum(2 if ord(ch) > 0xFF else 1 for ch in str(value))


class Sheet:
    """1枚のワークシートのラッパー。ExcelFile.sheet() から取得する。

    ここにないシート操作は .ws から生の openpyxl Worksheet を使える。
    """

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws  # 生の openpyxl Worksheet（高度な操作用に公開）

    # ------------------------------------------------------------ セル参照
    def __getitem__(self, ref: str):
        """セル参照で値を読む（例: s["A1"]）。"""
        return self.ws[ref].value

    def __setitem__(self, ref: str, value) -> None:
        """セル参照で値を書く（例: s["A1"] = "タイトル"）。"""
        self.ws[ref] = value

    # ------------------------------------------------------------ 行の書き込み
    def write_row(self, row: int, values: list, start_col: int = 1) -> None:
        """1行に値を横並びで書き込む。

        Args:
            row: 行番号（1始まり）。
            values: 書き込む値のリスト（左から順に並ぶ）。
            start_col: 開始列番号（1始まり。デフォルト: A列から）。
        """
        for i, value in enumerate(values):
            self.ws.cell(row=int(row), column=start_col + i).value = value

    def write_rows(self, start_row: int, rows: list[list], start_col: int = 1) -> None:
        """複数行をまとめて書き込む。

        Args:
            start_row: 開始行番号（1始まり）。
            rows: 行のリスト（値のリストのリスト）。
            start_col: 開始列番号（1始まり）。
        """
        for i, values in enumerate(rows):
            self.write_row(int(start_row) + i, values, start_col)

    def append_row(self, values: list) -> None:
        """最終行の下に1行追記する（空シートなら1行目に書く）。"""
        self.write_row(self.last_row + 1 if not self.is_empty else 1, values)

    def write_table(
        self, rows: list[dict], start_row: int = 1, headers: list[str] | None = None
    ) -> None:
        """ヘッダー行 + データ行をまとめて書き込む（CSV → Excel レポート等）。

        CsvReader.rows() や read_rows_as_dicts() の結果をそのまま渡せる。

        Args:
            rows: 辞書のリスト（キーが列名になる）。
            start_row: ヘッダー行の行番号（1始まり）。
            headers: 列の並び順。省略すると最初の行のキー順。
        """
        if not rows:
            return
        headers = headers or list(rows[0].keys())
        self.write_row(int(start_row), headers)
        for i, row in enumerate(rows, start=int(start_row) + 1):
            self.write_row(i, [row.get(h, "") for h in headers])

    # ------------------------------------------------------------ 見た目の調整
    def auto_width(self, min_width: float = 8, max_width: float = 60) -> None:
        """全列の幅を内容に合わせて調整する（全角文字は2文字ぶんで計算）。

        Args:
            min_width: 最小の列幅（内容が短くても これより狭くしない）。
            max_width: 最大の列幅（長文があっても これより広げない）。
        """
        for col_num, column in enumerate(self.ws.iter_cols(), start=1):
            width = max(
                (_display_width(cell.value) for cell in column if cell.value is not None),
                default=0,
            )
            letter = get_column_letter(col_num)
            self.ws.column_dimensions[letter].width = min(max(width + 2, min_width), max_width)

    def freeze_header(self, rows: int = 1) -> None:
        """ヘッダー行を固定する（スクロールしても見出しが消えない）。

        Args:
            rows: 固定する行数（デフォルト: 1行目のみ）。
        """
        self.ws.freeze_panes = f"A{int(rows) + 1}"

    # ------------------------------------------------------------ 状態の取得
    @property
    def last_row(self) -> int:
        """データがある最終行の番号（1始まり）。空シートでも 1 が返る点に注意。"""
        return self.ws.max_row

    @property
    def is_empty(self) -> bool:
        """シートに値が1つもないか返す。"""
        if self.ws.max_row > 1 or self.ws.max_column > 1:
            return False
        return self.ws.cell(row=1, column=1).value is None
