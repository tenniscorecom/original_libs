"""
excel/handler.py — Excel 読み書きユーティリティ（openpyxl ベース）

ExcelFile クラスを通じて Excel ファイルの読み書きを行う。
数式の計算結果が必要な場合や VBA マクロを実行する場合は、
内部で自動的に win32com（pywin32）にフォールバックする。

使い方:
    from comken.excel import ExcelFile

    with ExcelFile("data.xlsx") as f:
        rows = f.read_rows("Sheet1")
        f.write_cell("Sheet1", row=2, col=1, value="値")
        f.save()
"""

import logging
import shutil
import tempfile
from collections.abc import Generator
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..exceptions import ExcelError, SheetNotFoundError, _warn_coerce
from ..runtime import dry_run_log, is_dry_run
from ..utils.data import col_to_num
from ..utils.timer import measure
from .sheet import Sheet

logger = logging.getLogger(__name__)


class ExcelFile:
    """openpyxl ワークブックのラッパー。with 文で確実に閉じられる。

    通常の読み書きは openpyxl を使う。
    数式の計算結果が必要な場合は read_computed_rows()、
    VBA マクロを実行する場合は run_macro() を使う（どちらも win32com に自動切替）。

    使い方:
        # 読み取り
        with ExcelFile("data.xlsx") as f:
            rows = f.read_rows("Sheet1")
            # → [(値, 値, ...), ...]

            rows = f.read_rows_as_dicts("Sheet1")
            # → [{"列名": 値, ...}, ...]

        # ヘッダー行がない場合は __init__ で列名を渡す
        with ExcelFile("data.xlsx", headers=["注文番号", "金額", "担当者"]) as f:
            rows = f.read_rows_as_dicts("Sheet1")
            # → 1行目からデータとして読み込む

        # 数式の計算結果を読む（openpyxl → win32com 自動フォールバック）
        with ExcelFile("data.xlsx") as f:
            rows = f.read_computed_rows("Sheet1")

        # 書き込み
        with ExcelFile("data.xlsx") as f:
            f.write_cell("Sheet1", row=2, col=1, value="新しい値")
            f.save()

        # 別名で保存
        with ExcelFile("template.xlsx") as f:
            f.write_cell("Sheet1", row=2, col=1, value="値")
            f.save("output.xlsx")

        # VBA マクロの実行（常に win32com を使用）
        with ExcelFile("data.xlsm") as f:
            f.run_macro("Module1.UpdateData")
    """

    def __init__(
        self,
        path: str | Path,
        data_only: bool = False,
        read_only: bool = False,
        local_copy_threshold_mb: float = 10,
        headers: list[str] | None = None,
    ) -> None:
        """
        Args:
            path: Excel ファイルのパス。
            data_only: True にすると数式セルのキャッシュ値を読む（read_computed_rows 推奨）。
            read_only: True にすると読み取り専用で開く（大きなファイルで高速化）。
            local_copy_threshold_mb: この MB 以上のファイルはローカルにコピーしてから開く。
                NAS・ネットワークドライブのファイルが遅い・不安定な場合に有効。
                0 を指定するとローカルコピーを無効化できる。
            headers: ヘッダー行がない Excel の場合に、列名のリストをここで付ける。
                     指定すると read_rows_as_dicts() は全行をデータとして読む。
                     例: ExcelFile("data.xlsx", headers=["注文番号", "金額", "担当者"])
        """
        # save() の保存先は常に元のファイル（ローカルコピーに保存すると close で消えてしまう）
        self._original_path = Path(path)
        src = self._original_path

        # 素の FileNotFoundError ではなく、対処法つきの ExcelError にする
        if not src.exists():
            raise ExcelError(ExcelError.MSG_FILE_NOT_FOUND.format(path=src))

        # ── NAS・ネットワークファイルのローカルコピー ──────────────────
        # 社内ルールでローカルへのコピーが不可の場合は、
        # このブロックを丸ごと削除し「self._tmp = None」だけ残す。
        # close() 内の対応する削除ブロックも併せて削除すること。
        self._tmp = None
        if local_copy_threshold_mb and src.stat().st_size > local_copy_threshold_mb * 1024 * 1024:
            tmp = tempfile.NamedTemporaryFile(suffix=src.suffix, delete=False)
            self._tmp = Path(tmp.name)
            tmp.close()
            shutil.copy2(src, self._tmp)
            src = self._tmp
        # ────────────────────────────────────────────────────────────────

        self._path = src
        self._headers = headers
        # マクロ入りブック（.xlsm/.xlsb）は keep_vba=True で開かないと save() で VBA が消える
        keep_vba = self._original_path.suffix.lower() in (".xlsm", ".xlsb", ".xltm")
        self._wb: Workbook = load_workbook(
            self._path, data_only=data_only, read_only=read_only, keep_vba=keep_vba
        )

    @classmethod
    def create(cls, path: str | Path, sheet_name: str = "Sheet1") -> "ExcelFile":
        """新規ブックを作る（ファイルはまだ作られず、save() で path に保存される）。

        使い方:
            rows = CsvReader("data.csv").rows()
            with ExcelFile.create(r"C:\\作業\\report.xlsx") as f:
                s = f.sheet("Sheet1")
                s.write_table(rows)
                s.auto_width()
                f.save()

        Args:
            path: save() で保存されるパス。親フォルダがなければ保存時に自動作成される。
            sheet_name: 最初のシートの名前（デフォルト: "Sheet1"）。
        """
        instance = cls.__new__(cls)
        instance._original_path = Path(path)
        instance._tmp = None
        instance._path = instance._original_path
        instance._headers = None
        instance._wb = Workbook()
        instance._wb.active.title = sheet_name
        return instance

    def __enter__(self) -> "ExcelFile":
        return self

    def __exit__(self, *args) -> None:
        self.close()

    def sheet(self, name: str) -> Sheet:
        """シートの高レベルラッパーを返す（シート単位でセル・行を書き込む）。

        使い方:
            with ExcelFile("report.xlsx") as f:
                s = f.sheet("Sheet1")
                s["A1"] = "タイトル"
                s.write_row(3, ["日付", "金額"])
                s.auto_width()
                s.freeze_header()
                f.save()

        Args:
            name: シート名。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
        """
        return Sheet(self._sheet(name))

    def _sheet(self, name: str) -> Worksheet:
        """シートオブジェクトを返す。

        Args:
            name: シート名。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
        """
        name = _warn_coerce(name, str, "sheet_name", stacklevel=3)
        if name not in self._wb.sheetnames:
            raise SheetNotFoundError(
                SheetNotFoundError.MSG.format(name=name, sheets=self._wb.sheetnames)
            )
        return self._wb[name]

    @measure
    def read_rows(self, sheet_name: str, min_row: int = 2) -> list[tuple]:
        """指定シートの行データをタプルのリストで返す。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Returns:
            各行を値のタプルにしたリスト。
        """
        return list(self._sheet(sheet_name).iter_rows(min_row=int(min_row), values_only=True))

    @measure
    def read_rows_as_dicts(self, sheet_name: str, header_row: int = 1) -> list[dict]:
        """ヘッダー行をキーとした辞書のリストで返す。

        ヘッダー行がないファイルは ExcelFile(path, headers=[...]) で列名を指定すること。

        Args:
            sheet_name: シート名。
            header_row: ヘッダーが存在する行番号（デフォルト: 1）。
                        __init__ で headers を指定した場合は無視される。

        Returns:
            [{"列名": 値, ...}, ...] の形式のリスト。全セルが空の行は除外される。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
            ExcelError: ヘッダー行に空のセルがある場合（headers 未指定時のみ）、
                        または headers の列数がシートの列数より少ない場合。
        """
        ws = self._sheet(sheet_name)
        if self._headers is not None:
            all_rows = list(ws.iter_rows(min_row=1, values_only=True))
            if all_rows and len(all_rows[0]) > len(self._headers):
                raise ExcelError(
                    ExcelError.MSG_HEADERS_TOO_FEW.format(
                        expected=len(self._headers), actual=len(all_rows[0])
                    )
                )
            return [
                dict(zip(self._headers, row))
                for row in all_rows
                if not all(c is None for c in row)
            ]
        all_rows = list(ws.iter_rows(min_row=int(header_row), values_only=True))
        if not all_rows:
            return []
        file_headers = all_rows[0]
        if all(h is None for h in file_headers):
            return []
        none_cols = [i + 1 for i, h in enumerate(file_headers) if h is None]
        if none_cols:
            raise ExcelError(ExcelError.MSG_HEADER_NONE.format(cols=none_cols))
        return [dict(zip(file_headers, row)) for row in all_rows[1:]]

    def iter_rows(
        self, sheet_name: str, min_row: int = 2
    ) -> Generator[tuple[Any, ...], None, None]:
        """大量データ向け。行をジェネレータで1行ずつ返す（メモリ効率優先）。

        read_rows はファイル全体をメモリに乗せるため、数万行以上のファイルでは
        この メソッドを使って1行ずつ処理する。

        複数ファイルを同時に処理する場合（目安: 10ファイル以上）は
        concurrent.futures.ThreadPoolExecutor を使うとさらに高速化できる。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Yields:
            各行の値のタプル。
        """
        self._sheet(sheet_name)  # シート名の存在チェック（間違いを分かりやすいエラーにする）
        wb = load_workbook(self._path, data_only=True, read_only=True)
        try:
            for row in wb[str(sheet_name)].iter_rows(min_row=int(min_row), values_only=True):
                yield row
        finally:
            wb.close()

    @measure
    def read_computed_rows(self, sheet_name: str, min_row: int = 2) -> list[tuple]:
        """数式の計算結果を含む行を読む。

        まず openpyxl のキャッシュ値（data_only=True）で読もうとする。
        キャッシュが古いか数式文字列（=...）が残っている場合は
        win32com（pywin32）にフォールバックして Excel を起動して取得する。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Returns:
            各行を値のタプルにしたリスト。数式は計算後の値になっている。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
        """
        self._sheet(sheet_name)  # シート名の存在チェック（間違いを分かりやすいエラーにする）
        try:
            wb = load_workbook(self._path, data_only=True, read_only=True)
            rows = list(wb[str(sheet_name)].iter_rows(min_row=int(min_row), values_only=True))
            wb.close()
            has_formula = any(
                isinstance(cell, str) and cell.startswith("=") for row in rows for cell in row
            )
            if not has_formula:
                return rows
        except Exception as e:
            logger.debug("openpyxl での読み込みに失敗（%s）。win32com にフォールバックします", e)

        from ..windows.handler import ExcelComHandler

        with ExcelComHandler(self._path) as com:
            return com.read_rows(sheet_name, min_row)

    def write_cell(self, sheet_name: str, row: int, col: int, value) -> None:
        """セルに値を書き込む。

        Args:
            sheet_name: シート名。
            row: 行番号（1始まり）。
            col: 列番号（1始まり。A列=1、B列=2、…）。
            value: 書き込む値。
        """
        self._sheet(sheet_name).cell(row=int(row), column=int(col)).value = value

    @measure
    def transfer_by_key(
        self,
        sheet_name: str,
        key_col: int | str,
        lookup: dict[str, dict],
        column_mapping: dict[str, str],
        start_row: int = 2,
    ) -> int:
        """キー列の値で lookup を引き、一致した行に値を転記する（XLOOKUP 的転記）。

        ExcelComHandler.transfer_by_key の openpyxl 版。
        Excel を起動しないため数万行でも速い。数式の再計算が必要な場合だけ COM 版を使う。
        書き込み後は save() を忘れずに呼ぶこと。

        使い方:
            lookup = CsvReader("data.csv").index("注文番号")
            mapping = {"A": "顧客名", "B": "金額"}  # 列レター → lookup の列名

            with ExcelFile("data.xlsx") as f:
                matched = f.transfer_by_key("T_data", key_col="Q",
                                            lookup=lookup, column_mapping=mapping)
                f.save()

        Args:
            sheet_name: シート名。
            key_col: キー列。列レター（"Q"）または列番号（17）で指定する。
            lookup: {キーの値: {列名: 値}} の辞書。CsvReader.index() 等で作る。
            column_mapping: {列レター: lookup の列名} の辞書。
            start_row: 転記を始める行番号（デフォルト: 2。1行目はヘッダー想定）。

        Returns:
            転記した行数。

        Raises:
            SheetNotFoundError: 指定したシートが存在しない場合。
        """
        key_col_num = col_to_num(key_col) if isinstance(key_col, str) else int(key_col)
        mapping = {col_to_num(letter): name for letter, name in column_mapping.items()}

        ws = self._sheet(sheet_name)
        last_row = ws.max_row
        logger.info("シート「%s」: 最終行 %d行", sheet_name, last_row)

        matched = 0
        for row in range(int(start_row), last_row + 1):
            key_value = ws.cell(row=row, column=key_col_num).value
            if key_value is None or str(key_value).strip() == "":
                continue

            # 数値セルが float で入っていると "1001.0" になってしまうため、
            # 整数値なら int を経由して "1001" に揃える（CSV 側の文字列と一致させる）
            if isinstance(key_value, float) and key_value.is_integer():
                key_value = int(key_value)

            lookup_row = lookup.get(str(key_value).strip())
            if lookup_row is None:
                logger.debug("%d行目: キー「%s」が lookup に存在しません", row, key_value)
                continue

            for col_num, name in mapping.items():
                ws.cell(row=row, column=col_num).value = lookup_row.get(name, "")
            logger.debug("%d行目: 転記完了（キー: %s）", row, key_value)
            matched += 1

        logger.info("転記完了: %d件一致（シート: %s）", matched, sheet_name)
        return matched

    @measure
    def save(self, path: str | Path | None = None) -> None:
        """ファイルを保存する。

        ローカルコピーで開いている場合も、省略時の保存先は元のファイル
        （一時コピーに保存すると close() でコピーごと消えてしまうため）。

        Args:
            path: 保存先のパス。省略すると開いた元のファイルに上書き保存する。
        """
        save_path = Path(path) if path else self._original_path
        if is_dry_run():
            dry_run_log("Excel を保存: %s", save_path)
            return
        save_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(save_path)

    def set_fill(self, sheet_name: str, row: int, col: int, color: str) -> None:
        """セルの背景色を設定する。

        よく使う色コード:
            "FFFF00" → 黄色
            "FF0000" → 赤
            "00FF00" → 緑
            "FFFFFF" → 白（色なし）

        使い方:
            with ExcelFile("data.xlsx") as f:
                f.set_fill("Sheet1", row=2, col=1, color="FFFF00")
                f.save()

        Args:
            sheet_name: シート名。
            row: 行番号（1始まり）。
            col: 列番号（1始まり）。
            color: 16進数カラーコード（"RRGGBB" 形式、# なし）。
        """
        color = _warn_coerce(color, str, "color", stacklevel=2)
        fill = PatternFill(fill_type="solid", fgColor=color)
        self._sheet(sheet_name).cell(row=int(row), column=int(col)).fill = fill

    def set_column_width(self, sheet_name: str, col: int, width: float) -> None:
        """列幅を設定する。

        Excel の列幅の目安: 標準フォント（11pt）で 1文字 ≈ 1。
        日本語文字は全角なので 2文字分として計算する（"山田太郎" = 8程度）。

        使い方:
            with ExcelFile("data.xlsx") as f:
                f.set_column_width("Sheet1", col=1, width=20)  # A列を幅20に
                f.save()

        Args:
            sheet_name: シート名。
            col: 列番号（1始まり。A列=1、B列=2、…）。
            width: 列幅（Excel の列幅単位）。
        """
        col_letter = get_column_letter(int(col))
        self._sheet(sheet_name).column_dimensions[col_letter].width = float(width)

    def set_number_format(self, sheet_name: str, row: int, col: int, fmt: str) -> None:
        """セルの数値フォーマットを設定する。

        よく使うフォーマット:
            "#,##0"          → 1,000（カンマ区切り整数）
            "#,##0.00"       → 1,000.00（小数2桁）
            "0%"             → 50%（パーセント）
            "yyyy/mm/dd"     → 2026/07/10（日付）
            "yyyy/mm/dd hh:mm" → 2026/07/10 09:00（日時）
            "@"              → 文字列として扱う

        使い方:
            with ExcelFile("data.xlsx") as f:
                f.set_number_format("Sheet1", row=2, col=3, fmt="#,##0")
                f.save()

        Args:
            sheet_name: シート名。
            row: 行番号（1始まり）。
            col: 列番号（1始まり）。
            fmt: Excel の書式文字列。
        """
        fmt = _warn_coerce(fmt, str, "fmt", stacklevel=2)
        self._sheet(sheet_name).cell(row=int(row), column=int(col)).number_format = fmt

    def set_bold(self, sheet_name: str, row: int, col: int, bold: bool = True) -> None:
        """セルの太字を設定する。

        使い方:
            with ExcelFile("data.xlsx") as f:
                f.set_bold("Sheet1", row=1, col=1)  # ヘッダーを太字に
                f.save()

        Args:
            sheet_name: シート名。
            row: 行番号（1始まり）。
            col: 列番号（1始まり）。
            bold: True で太字、False で解除。
        """
        cell = self._sheet(sheet_name).cell(row=int(row), column=int(col))
        cell.font = Font(bold=bool(bold))

    def run_macro(self, macro_name: str, save: bool = True) -> None:
        """VBA マクロを実行する。内部で win32com（pywin32）を使用する。

        COM は保存せずに閉じる仕様のため、save=True（デフォルト）で実行後に
        元ファイルへ保存する。マクロがブックを変更しても保存しないと結果は破棄される。

        WARNING: このメソッドは COM で元ファイルを直接編集する。openpyxl 側
            （write_cell 等）の未保存の変更とは独立で、run_macro の後に f.save() を
            呼ぶと openpyxl の内容で上書きされマクロの結果が消える。
            マクロと openpyxl 書き込みを混在させないこと。

        Args:
            macro_name: 実行するマクロ名。"モジュール名.プロシージャ名" の形式で指定する。
                        例: "Module1.UpdateData"
            save: True（デフォルト）ならマクロ実行後に元ファイルへ保存する。
        """
        from ..windows.handler import ExcelComHandler

        # local_copy の一時コピーではなく元ファイルに対して実行・保存する
        with ExcelComHandler(self._original_path) as com:
            com.run_macro(macro_name)
            if save:
                com.save()

    def close(self) -> None:
        """ワークブックを閉じる。with 文を使う場合は自動で呼ばれる。"""
        self._wb.close()

        # ── ローカルコピーの後処理（__init__ の対応ブロックと一緒に削除）──
        if self._tmp:
            self._tmp.unlink(missing_ok=True)
        # ──────────────────────────────────────────────────────────────────
