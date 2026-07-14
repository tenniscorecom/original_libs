"""
ExcelFile クラスのテスト。

実行方法:
    リポジトリのルートで python -m pytest tests/ -v
"""

import pytest
from openpyxl import Workbook, load_workbook

from comken.excel.handler import ExcelFile
from comken.exceptions import ExcelError, SheetNotFoundError


@pytest.fixture
def excel_with_header(tmp_path):
    """1行目がヘッダーの Excel ファイルを作成して返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["注文番号", "金額", "担当者"])
    ws.append(["A001", 1000, "山田"])
    ws.append(["A002", 2000, "佐藤"])
    path = tmp_path / "data.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def excel_no_header(tmp_path):
    """ヘッダー行なし（全行データ）の Excel ファイルを作成して返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["A001", 1000, "山田"])
    ws.append(["A002", 2000, "佐藤"])
    path = tmp_path / "no_header.xlsx"
    wb.save(path)
    return path


@pytest.fixture
def excel_header_row2(tmp_path):
    """2行目がヘッダーの Excel ファイルを作成して返す。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["（この行は無視）", None, None])
    ws.append(["注文番号", "金額", "担当者"])
    ws.append(["A001", 1000, "山田"])
    path = tmp_path / "header_row2.xlsx"
    wb.save(path)
    return path


class TestOpen:
    def test_missing_file_raises_excel_error(self, tmp_path):
        """存在しないファイルは素の FileNotFoundError ではなく ExcelError になる。"""
        with pytest.raises(ExcelError, match="見つかりません"):
            ExcelFile(tmp_path / "no_such.xlsx")


class TestReadRowsAsDicts:
    """read_rows_as_dicts() の基本動作テスト。"""

    def test_reads_all_rows(self, excel_with_header):
        """1行目ヘッダーの場合に全データ行を辞書で返すことを確認する。"""
        with ExcelFile(excel_with_header) as f:
            rows = f.read_rows_as_dicts("Sheet1")
        assert len(rows) == 2
        assert rows[0] == {"注文番号": "A001", "金額": 1000, "担当者": "山田"}

    def test_header_row_parameter(self, excel_header_row2):
        """header_row=2 の場合に2行目をヘッダーとして読むことを確認する。"""
        with ExcelFile(excel_header_row2) as f:
            rows = f.read_rows_as_dicts("Sheet1", header_row=2)
        assert len(rows) == 1
        assert rows[0]["注文番号"] == "A001"

    def test_empty_sheet_returns_empty_list(self, tmp_path):
        """空のシートは空リストを返すことを確認する。"""
        wb = Workbook()
        wb.active.title = "Sheet1"
        path = tmp_path / "empty.xlsx"
        wb.save(path)
        with ExcelFile(path) as f:
            rows = f.read_rows_as_dicts("Sheet1")
        assert rows == []

    def test_raises_on_missing_sheet(self, excel_with_header):
        """存在しないシートを指定すると SheetNotFoundError になることを確認する。"""
        with ExcelFile(excel_with_header) as f:
            with pytest.raises(SheetNotFoundError):
                f.read_rows_as_dicts("存在しないシート")

    def test_raises_on_none_header_cell(self, tmp_path):
        """ヘッダー行に空のセルがある場合は ExcelError になることを確認する。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["注文番号", None, "担当者"])
        ws.append(["A001", 1000, "山田"])
        path = tmp_path / "none_header.xlsx"
        wb.save(path)
        with ExcelFile(path) as f:
            with pytest.raises(ExcelError, match="空のセル"):
                f.read_rows_as_dicts("Sheet1")


class TestReadRowsAsDictsWithHeaders:
    """ExcelFile(path, headers=...) のテスト（ヘッダー行なしファイル）。"""

    def test_reads_headerless_file(self, excel_no_header):
        """__init__ で headers を指定すると全行をデータとして読めることを確認する。"""
        with ExcelFile(excel_no_header, headers=["注文番号", "金額", "担当者"]) as f:
            rows = f.read_rows_as_dicts("Sheet1")
        assert len(rows) == 2
        assert rows[0] == {"注文番号": "A001", "金額": 1000, "担当者": "山田"}
        assert rows[1]["注文番号"] == "A002"

    def test_headers_overrides_file_headers(self, excel_with_header):
        """__init__ で headers 指定時はファイルの1行目もデータとして読むことを確認する。

        （ヘッダーありファイルに headers を渡すと、ヘッダー行もデータになる）
        """
        with ExcelFile(excel_with_header, headers=["C1", "C2", "C3"]) as f:
            rows = f.read_rows_as_dicts("Sheet1")
        assert len(rows) == 3  # ヘッダー行を含む全3行
        assert rows[0] == {"C1": "注文番号", "C2": "金額", "C3": "担当者"}

    def test_headers_applies_to_all_sheets(self, tmp_path):
        """__init__ で headers を指定すると全シートに適用されることを確認する。"""
        wb = Workbook()
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1.append(["A001", 1000])
        ws2 = wb.create_sheet("Sheet2")
        ws2.append(["B001", 2000])
        path = tmp_path / "multi.xlsx"
        wb.save(path)

        with ExcelFile(path, headers=["注文番号", "金額"]) as f:
            rows1 = f.read_rows_as_dicts("Sheet1")
            rows2 = f.read_rows_as_dicts("Sheet2")
        assert rows1[0]["注文番号"] == "A001"
        assert rows2[0]["注文番号"] == "B001"

    def test_headers_too_few_raises(self, excel_no_header):
        """headers の列数がシートの列数より少ないとエラーになることを確認する。

        （zip が黙って列を落とすとデータ欠損に気づけないため）
        """
        with ExcelFile(excel_no_header, headers=["注文番号", "金額"]) as f:  # 実際は3列
            with pytest.raises(ExcelError, match="列数"):
                f.read_rows_as_dicts("Sheet1")

    def test_headers_with_empty_sheet_returns_empty(self, tmp_path):
        """headers 指定でも空シートは空リストを返すことを確認する（偽の1行を返さない）。"""
        wb = Workbook()
        wb.active.title = "Sheet1"
        path = tmp_path / "empty.xlsx"
        wb.save(path)

        with ExcelFile(path, headers=["注文番号", "金額"]) as f:
            rows = f.read_rows_as_dicts("Sheet1")
        assert rows == []

    def test_headers_skips_blank_rows(self, tmp_path):
        """全セルが空の行は結果に含まれないことを確認する。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["A001", 1000])
        ws.append([None, None])  # 空行
        ws.append(["A002", 2000])
        path = tmp_path / "blank_row.xlsx"
        wb.save(path)

        with ExcelFile(path, headers=["注文番号", "金額"]) as f:
            rows = f.read_rows_as_dicts("Sheet1")
        assert len(rows) == 2
        assert rows[1]["注文番号"] == "A002"


class TestTransferByKey:
    """ExcelFile.transfer_by_key（openpyxl 版のキー突合転記）のテスト。"""

    @pytest.fixture
    def transfer_excel(self, tmp_path):
        """転記先の Excel（キー列 A、転記先列 B・C）を作成して返す。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "T_data"
        ws.append(["注文番号", "顧客名", "金額"])
        ws.append(["A001", None, None])
        ws.append(["A002", None, None])
        ws.append(["Z999", None, None])  # lookup に存在しないキー
        path = tmp_path / "transfer.xlsx"
        wb.save(path)
        return path

    def test_transfers_matching_rows(self, transfer_excel):
        """キーが一致した行に値が転記され、件数が返ることを確認する。"""
        lookup = {
            "A001": {"顧客名": "株式会社A", "金額": "1000"},
            "A002": {"顧客名": "株式会社B", "金額": "2000"},
        }

        with ExcelFile(transfer_excel) as f:
            matched = f.transfer_by_key(
                "T_data", key_col="A", lookup=lookup, column_mapping={"B": "顧客名", "C": "金額"}
            )
            f.save()

        assert matched == 2
        wb = load_workbook(transfer_excel)
        ws = wb["T_data"]
        assert ws.cell(row=2, column=2).value == "株式会社A"
        assert ws.cell(row=3, column=3).value == "2000"
        wb.close()

    def test_skips_missing_keys(self, transfer_excel):
        """lookup にないキーの行は転記されずスキップされることを確認する。"""
        lookup = {"A001": {"顧客名": "株式会社A"}}

        with ExcelFile(transfer_excel) as f:
            matched = f.transfer_by_key(
                "T_data", key_col="A", lookup=lookup, column_mapping={"B": "顧客名"}
            )
            f.save()

        assert matched == 1
        wb = load_workbook(transfer_excel)
        ws = wb["T_data"]
        assert ws.cell(row=4, column=2).value is None  # Z999 の行は未転記
        wb.close()

    def test_float_integer_key_matches_csv_string(self, tmp_path):
        """数値キー（1001.0）が CSV 側の文字列キー "1001" と突合できることを確認する。"""
        wb = Workbook()
        ws = wb.active
        ws.title = "T_data"
        ws.append(["注文番号", "顧客名"])
        ws.append([1001.0, None])
        path = tmp_path / "float_key.xlsx"
        wb.save(path)

        lookup = {"1001": {"顧客名": "株式会社C"}}

        with ExcelFile(path) as f:
            matched = f.transfer_by_key(
                "T_data", key_col="A", lookup=lookup, column_mapping={"B": "顧客名"}
            )
            f.save()

        assert matched == 1
        wb = load_workbook(path)
        assert wb["T_data"].cell(row=2, column=2).value == "株式会社C"
        wb.close()

    def test_key_col_accepts_column_number(self, transfer_excel):
        """key_col を列レターではなく列番号（1）で指定できることを確認する。"""
        lookup = {"A001": {"顧客名": "株式会社A"}}

        with ExcelFile(transfer_excel) as f:
            matched = f.transfer_by_key(
                "T_data", key_col=1, lookup=lookup, column_mapping={"B": "顧客名"}
            )

        assert matched == 1

    def test_raises_on_missing_sheet(self, transfer_excel):
        """存在しないシートを指定すると SheetNotFoundError になることを確認する。"""
        with ExcelFile(transfer_excel) as f:
            with pytest.raises(SheetNotFoundError):
                f.transfer_by_key("存在しない", key_col="A", lookup={}, column_mapping={})


class TestSheetWrapper:
    """Sheet（シート単位の高レベルラッパー）のテスト。"""

    def test_create_and_cell_access(self, tmp_path):
        """ExcelFile.create で新規ブックを作り、セル参照で読み書きできることを確認する。"""
        path = tmp_path / "new.xlsx"
        with ExcelFile.create(path) as f:
            s = f.sheet("Sheet1")
            s["A1"] = "タイトル"

            assert s["A1"] == "タイトル"
            f.save()

        assert path.exists()

    def test_write_row_and_rows(self, tmp_path):
        """write_row / write_rows で横並びに書き込まれることを確認する。"""
        path = tmp_path / "rows.xlsx"
        with ExcelFile.create(path) as f:
            s = f.sheet("Sheet1")
            s.write_row(1, ["日付", "金額"])
            s.write_rows(2, [["7/1", 100], ["7/2", 200]])
            f.save()

        with ExcelFile(path) as f:
            rows = f.read_rows_as_dicts("Sheet1")
            assert rows == [{"日付": "7/1", "金額": 100}, {"日付": "7/2", "金額": 200}]

    def test_append_row_on_empty_sheet_starts_at_row1(self, tmp_path):
        """空シートへの append_row は1行目から書かれることを確認する（2行目から始まらない）。"""
        with ExcelFile.create(tmp_path / "a.xlsx") as f:
            s = f.sheet("Sheet1")
            assert s.is_empty

            s.append_row(["ヘッダー"])
            s.append_row(["データ"])

            assert s["A1"] == "ヘッダー"
            assert s["A2"] == "データ"
            assert not s.is_empty

    def test_write_table_writes_header_and_rows(self, tmp_path):
        """write_table で辞書のリストがヘッダー付きで書かれることを確認する。"""
        rows = [{"注文番号": "A001", "金額": 1000}, {"注文番号": "A002", "金額": 2000}]
        path = tmp_path / "table.xlsx"
        with ExcelFile.create(path) as f:
            f.sheet("Sheet1").write_table(rows)
            f.save()

        with ExcelFile(path) as f:
            assert f.read_rows_as_dicts("Sheet1") == rows

    def test_write_table_respects_header_order(self, tmp_path):
        """headers 指定で列の並び順を制御できることを確認する。"""
        rows = [{"金額": 1000, "注文番号": "A001"}]
        with ExcelFile.create(tmp_path / "t.xlsx") as f:
            s = f.sheet("Sheet1")
            s.write_table(rows, headers=["注文番号", "金額"])

            assert s["A1"] == "注文番号"
            assert s["B1"] == "金額"

    def test_auto_width_considers_japanese(self, tmp_path):
        """auto_width で全角文字が2文字ぶんとして幅計算されることを確認する。"""
        with ExcelFile.create(tmp_path / "w.xlsx") as f:
            s = f.sheet("Sheet1")
            s["A1"] = "日本語のタイトル"  # 8文字 → 表示幅16
            s["B1"] = "abc"

            s.auto_width()

            assert s.ws.column_dimensions["A"].width >= 16
            assert s.ws.column_dimensions["B"].width == 8  # min_width

    def test_auto_width_caps_at_max(self, tmp_path):
        """長文があっても max_width を超えないことを確認する。"""
        with ExcelFile.create(tmp_path / "w.xlsx") as f:
            s = f.sheet("Sheet1")
            s["A1"] = "あ" * 100

            s.auto_width(max_width=60)

            assert s.ws.column_dimensions["A"].width == 60

    def test_freeze_header(self, tmp_path):
        """freeze_header で1行目（指定行数）が固定されることを確認する。"""
        with ExcelFile.create(tmp_path / "f.xlsx") as f:
            s = f.sheet("Sheet1")
            s.freeze_header()
            assert s.ws.freeze_panes == "A2"

            s.freeze_header(rows=2)
            assert s.ws.freeze_panes == "A3"

    def test_sheet_raises_on_missing_sheet(self, tmp_path):
        """存在しないシート名は SheetNotFoundError になることを確認する。"""
        with ExcelFile.create(tmp_path / "e.xlsx") as f:
            with pytest.raises(SheetNotFoundError):
                f.sheet("存在しないシート")
