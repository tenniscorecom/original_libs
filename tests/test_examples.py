"""examples/ のオフラインサンプルが実際に動くことを確認するテスト。

外部システム（Excel COM・ブラウザ・Salesforce）を必要としない3本を、
出力先を tmp_path に差し替えて main() まで通し、成果物ができることを検証する。
README・ドキュメントが「そのまま動く」と案内している主張をテストで担保する。
"""

from openpyxl import load_workbook


class TestCsvToExcelReport:
    def test_creates_report(self, tmp_path, monkeypatch):
        """CSV を読んで Excel レポートを作る例が xlsx を出力する。"""
        from examples.csv_to_excel_report import run

        monkeypatch.setattr(run, "OUTPUT_FOLDER", tmp_path)
        run.main()

        outputs = list(tmp_path.glob("*.xlsx"))
        assert len(outputs) == 1
        # 合計行まで書けている（ヘッダー + データ3件 + 合計 = 5行以上）
        ws = load_workbook(outputs[0]).active
        assert ws.max_row >= 5
        assert ws.cell(row=ws.max_row, column=1).value == "合計"


class TestExcelKeyTransfer:
    def test_transfers_matched_rows(self, tmp_path, monkeypatch):
        """キー突合転記の例が、マスタにあるキーだけ転記する。"""
        from examples.excel_key_transfer import run

        monkeypatch.setattr(run, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(run, "MASTER_CSV", tmp_path / "master.csv")
        monkeypatch.setattr(run, "INVOICE_XLSX", tmp_path / "invoice.xlsx")
        run.main()

        ws = load_workbook(tmp_path / "invoice.xlsx").active
        rows = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True)}
        # A001 は転記され顧客名が入る、Z999 はマスタにないので空のまま
        assert rows["A001"][1] == "株式会社アルファ"
        assert rows["Z999"][1] in (None, "")


class TestCsvDiffReport:
    def test_detects_added_removed_changed(self, tmp_path, monkeypatch):
        """差分レポートの例が追加・削除・変更を検出して xlsx を出す。"""
        from examples.csv_diff_report import run

        monkeypatch.setattr(run, "OUTPUT_FOLDER", tmp_path)
        monkeypatch.setattr(run, "YESTERDAY_CSV", tmp_path / "yesterday.csv")
        monkeypatch.setattr(run, "TODAY_CSV", tmp_path / "today.csv")
        run.main()

        outputs = list(tmp_path.glob("*.xlsx"))
        assert len(outputs) == 1
        statuses = {
            row[0]
            for row in load_workbook(outputs[0]).active.iter_rows(min_row=2, values_only=True)
        }
        # 追加(004)・削除(003)・変更(002) がそれぞれ検出されている
        assert {"追加", "削除", "変更"} <= statuses
