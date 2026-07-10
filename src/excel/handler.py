"""
excel/handler.py — Excel 読み書きユーティリティ（openpyxl ベース）

ExcelFile クラスを通じて Excel ファイルの読み書きを行う。
数式の計算結果が必要な場合や VBA マクロを実行する場合は、
内部で自動的に win32com（pywin32）にフォールバックする。

使い方:
    from src.excel.handler import ExcelFile

    with ExcelFile("data.xlsx") as f:
        rows = f.read_rows("Sheet1")
        f.write_cell("Sheet1", row=2, col=1, value="値")
        f.save()
"""

import shutil
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet.worksheet import Worksheet


class ExcelFile:
    """openpyxl ワークブックのラッパー。with 文で確実に閉じられる。

    通常の読み書きは openpyxl を使う。
    数式の計算結果が必要な場合は read_computed_rows()、
    VBA マクロを実行する場合は run_macro() を使う（どちらも win32com に自動切替）。

    使い方:
        # 読み取り
        with ExcelFile("data.xlsx") as f:
            rows = f.read_rows("Sheet1")
            # → [(値, 値, ...), ...]

            rows = f.read_rows_as_dicts("Sheet1")
            # → [{"列名": 値, ...}, ...]

        # 数式の計算結果を読む（openpyxl → win32com 自動フォールバック）
        with ExcelFile("data.xlsx") as f:
            rows = f.read_computed_rows("Sheet1")

        # 書き込み
        with ExcelFile("data.xlsx") as f:
            f.write_cell("Sheet1", row=2, col=1, value="新しい値")
            f.save()

        # 別名で保存
        with ExcelFile("template.xlsx") as f:
            f.write_cell("Sheet1", row=2, col=1, value="値")
            f.save("output.xlsx")

        # VBA マクロの実行（常に win32com を使用）
        with ExcelFile("data.xlsm") as f:
            f.run_macro("Module1.UpdateData")
    """

    def __init__(
        self,
        path: str | Path,
        data_only: bool = False,
        read_only: bool = False,
        local_copy_threshold_mb: int = 10,
    ) -> None:
        """
        Args:
            path: Excel ファイルのパス。
            data_only: True にすると数式セルのキャッシュ値を読む（read_computed_rows 推奨）。
            read_only: True にすると読み取り専用で開く（大きなファイルで高速化）。
            local_copy_threshold_mb: この MB 以上のファイルはローカルにコピーしてから開く。
                NAS・ネットワークドライブのファイルが遅い・不安定な場合に有効。
                0 を指定するとローカルコピーを無効化できる。
        """
        src = Path(path)

        # ── NAS・ネットワークファイルのローカルコピー ──────────────────
        # 社内ルールでローカルへのコピーが不可の場合は、
        # このブロックを丸ごと削除し「self._tmp = None」だけ残す。
        # close() 内の対応する削除ブロックも併せて削除すること。
        self._tmp = None
        if local_copy_threshold_mb and src.stat().st_size > local_copy_threshold_mb * 1024 * 1024:
            tmp = tempfile.NamedTemporaryFile(suffix=src.suffix, delete=False)
            self._tmp = Path(tmp.name)
            tmp.close()
            shutil.copy2(src, self._tmp)
            src = self._tmp
        # ────────────────────────────────────────────────────────────────

        self._path = src
        self._wb: Workbook = load_workbook(self._path, data_only=data_only, read_only=read_only)

    def __enter__(self) -> "ExcelFile":
        return self

    def __exit__(self, *args) -> None:
        self.close()

    def sheet(self, name: str) -> Worksheet:
        """シートオブジェクトを返す。

        Args:
            name: シート名。

        Raises:
            ValueError: 指定したシートが存在しない場合。
        """
        if name not in self._wb.sheetnames:
            raise ValueError(f"シートが見つかりません: {name}  存在するシート: {self._wb.sheetnames}")
        return self._wb[name]

    def read_rows(self, sheet_name: str, min_row: int = 2) -> list[tuple]:
        """指定シートの行データをタプルのリストで返す。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Returns:
            各行を値のタプルにしたリスト。
        """
        return list(self.sheet(sheet_name).iter_rows(min_row=min_row, values_only=True))

    def read_rows_as_dicts(self, sheet_name: str, header_row: int = 1) -> list[dict]:
        """ヘッダー行をキーとした辞書のリストで返す。

        Args:
            sheet_name: シート名。
            header_row: ヘッダーが存在する行番号（デフォルト: 1）。

        Returns:
            [{"列名": 値, ...}, ...] の形式のリスト。
        """
        ws = self.sheet(sheet_name)
        all_rows = list(ws.iter_rows(min_row=header_row, values_only=True))
        if not all_rows:
            return []
        headers = all_rows[0]
        return [dict(zip(headers, row)) for row in all_rows[1:]]

    def iter_rows(self, sheet_name: str, min_row: int = 2):
        """大量データ向け。行をジェネレータで1行ずつ返す（メモリ効率優先）。

        read_rows はファイル全体をメモリに乗せるため、数万行以上のファイルでは
        この メソッドを使って1行ずつ処理する。

        複数ファイルを同時に処理する場合（目安: 10ファイル以上）は
        concurrent.futures.ThreadPoolExecutor を使うとさらに高速化できる。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Yields:
            各行の値のタプル。
        """
        wb = load_workbook(self._path, data_only=True, read_only=True)
        try:
            for row in wb[sheet_name].iter_rows(min_row=min_row, values_only=True):
                yield row
        finally:
            wb.close()

    def read_computed_rows(self, sheet_name: str, min_row: int = 2) -> list[tuple]:
        """数式の計算結果を含む行を読む。

        まず openpyxl のキャッシュ値（data_only=True）で読もうとする。
        キャッシュが古いか数式文字列（=...）が残っている場合は
        win32com（pywin32）にフォールバックして Excel を起動して取得する。

        Args:
            sheet_name: シート名。
            min_row: 読み始める行番号（デフォルト: 2 でヘッダーをスキップ）。

        Returns:
            各行を値のタプルにしたリスト。数式は計算後の値になっている。
        """
        try:
            wb = load_workbook(self._path, data_only=True, read_only=True)
            rows = list(wb[sheet_name].iter_rows(min_row=min_row, values_only=True))
            wb.close()
            has_formula = any(
                isinstance(cell, str) and cell.startswith("=")
                for row in rows for cell in row
            )
            if not has_formula:
                return rows
        except Exception:
            pass

        from ..windows.handler import ExcelComHandler
        with ExcelComHandler(self._path) as com:
            return com.read_rows(sheet_name, min_row)

    def write_cell(self, sheet_name: str, row: int, col: int, value) -> None:
        """セルに値を書き込む。

        Args:
            sheet_name: シート名。
            row: 行番号（1始まり）。
            col: 列番号（1始まり。A列=1、B列=2、…）。
            value: 書き込む値。
        """
        self.sheet(sheet_name).cell(row=row, column=col).value = value

    def save(self, path: str | Path | None = None) -> None:
        """ファイルを保存する。

        Args:
            path: 保存先のパス。省略すると開いたファイルに上書き保存する。
        """
        save_path = Path(path) if path else self._path
        save_path.parent.mkdir(parents=True, exist_ok=True)
        self._wb.save(save_path)

    def set_fill(self, sheet_name: str, row: int, col: int, color: str) -> None:
        """セルの背景色を設定する。

        よく使う色コード:
            "FFFF00" → 黄色
            "FF0000" → 赤
            "00FF00" → 緑
            "FFFFFF" → 白（色なし）

        使い方:
            with ExcelFile("data.xlsx") as f:
                f.set_fill("Sheet1", row=2, col=1, color="FFFF00")
                f.save()

        Args:
            sheet_name: シート名。
            row: 行番号（1始まり）。
            col: 列番号（1始まり）。
            color: 16進数カラーコード（"RRGGBB" 形式、# なし）。
        """
        fill = PatternFill(fill_type="solid", fgColor=color)
        self.sheet(sheet_name).cell(row=row, column=col).fill = fill

    def run_macro(self, macro_name: str) -> None:
        """VBA マクロを実行する。内部で win32com（pywin32）を使用する。

        Args:
            macro_name: 実行するマクロ名。"モジュール名.プロシージャ名" の形式で指定する。
                        例: "Module1.UpdateData"
        """
        from ..windows.handler import ExcelComHandler
        with ExcelComHandler(self._path) as com:
            com.run_macro(macro_name)

    def close(self) -> None:
        """ワークブックを閉じる。with 文を使う場合は自動で呼ばれる。"""
        self._wb.close()

        # ── ローカルコピーの後処理（__init__ の対応ブロックと一緒に削除）──
        if self._tmp:
            self._tmp.unlink(missing_ok=True)
        # ──────────────────────────────────────────────────────────────────
